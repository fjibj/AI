# coding:utf-8
import datetime
import math
import re
from urllib import request, parse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 得到金字对应的陈列盘位置
def getJinPosition(hz):
    s = parse.quote(hz)
    url = 'http://www.zizaicn.com/weixin/xixifu_zzping_search_word.php?str=' + s
    f = request.urlopen(url)
    h = f.read().decode('utf-8')
    # print(h)
    m = re.search(r'查询结果：</strong><br>(.*):第(.*)行 (.*)列</div>', h)
    if m:
        return int(m.group(2)), int(m.group(3))
    else:
        return 1000, 1000


# 得到铅字对应的陈列盘位置
def getposition(hz):
    # 先从本地position_qianzi.csv中查找，如果没找到再从微信链接查询
    pqz = pd.read_csv('position_qianzi.csv')
    if(not pqz.loc[pqz['A']==hz].empty):
        print('get pos:'+ hz)
        return int(pqz.loc[pqz['A']==hz]['C']),int(pqz.loc[pqz['A']==hz]['D']),int(pqz.loc[pqz['A']==hz]['E'])
    else:
        s = parse.quote(hz)
        url = 'http://www.zizaicn.com/weixin/getwordpos_ver2.php?str=' + s
        f = request.urlopen(url)
        h = f.read().decode('utf-8')
        m = re.search(r'查询结果：</strong><br>(.*)<br></div>', h)
        m2 = re.match(r'(.*):(.*)盘 (.*)行 (.*)列', m.group(1))
        if m2:
            # 查询结果添加到position_qianzi.csv中
            pp = pd.DataFrame({'A':[hz], 'C':[int(m2.group(2))], 'D':[int(m2.group(3))], 'E':[int(m2.group(4))]})
            pp.to_csv('position_qianzi.csv', mode='a', index=False, header=False)
            print('write to csv:' + hz)
            return int(m2.group(2)), int(m2.group(3)), int(m2.group(4))
        else:
            return 1000, 1000, 1000


# 读金字.csv文件生成DataFrame
def sortJinPosition(csvfile):
    s = datetime.datetime.now()
    df = pd.read_csv(csvfile, header=None, names=['A', 'B'])
    # print(df['A'])
    df[['C', 'D']] = df.apply(lambda row: pd.Series(getJinPosition(row['A'])), axis=1)
    # print(df)
    # df.to_csv('sort1.csv', encoding='utf-8')
    df2 = df.sort_values(['C', 'D'], ascending=[True, True])
    df2.reset_index(drop=True, inplace=True)
    df2.to_csv('sort2.csv', encoding='utf-8')
    e = datetime.datetime.now()
    print('sort spended ' + str((e - s).seconds) + ' seconds.')
    return df2


# 读铅字.csv文件生成DataFrame
# isSum： True 进货量求和，false, 进货量取最大值
def sortposition(csvfile, isSum):
    s = datetime.datetime.now()
    df = pd.read_csv(csvfile, header=None, names=['A', 'B'])
    if isSum:
        df1 = df.groupby('A',as_index=False).sum() #重复汉字进货量求和
    else:
        #重复汉字取最大进货量
        df1 = df.sort_values(['B'], ascending=[False]).drop_duplicates(subset=['A'], keep='first')
    print(df1)
    df1[['C', 'D', 'E']] = df1.apply(lambda row: pd.Series(getposition(row['A'])), axis=1)
    # print(df)
    # df.to_csv('sort1.csv', encoding='utf-8')
    # df2 = df.sort_values(['C', 'D', 'E'], ascending=[True, True, True])
    # 重复汉字只保留进货量大的一条记录
    #df2 = df.sort_values(['C', 'D', 'E', 'B'], ascending=[True, True, True, False]).drop_duplicates(subset=['A'],
    #                                                                                                keep='first')
    df2 = df1.sort_values(['C', 'D', 'E'], ascending=[True, True, True])
    df2.reset_index(drop=True, inplace=True)
    df2.to_csv('sort2.csv', encoding='utf-8')
    e = datetime.datetime.now()
    print('sort spended ' + str((e - s).seconds) + ' seconds.')
    return df2


# 获得在Excel文档中的行列位置(行，列，页）注：页表示对应的标签名称最后的()中的数字，如果为1，则标签名称不含()
# index：在df中的序号
# sc：起始列
# ine：列间隔
# ec: 终止列（数字列）
# sl：起始行
# el：终止行
def getxlspos(index, sc, ine, ec, sl, el):
    row = index % (el - sl + 1) + sl
    col = chr(ord(sc) + (index // (el - sl + 1)) * ine % (ord(ec) - ord(sc) + 1))
    page = 1 + (index // ((ord(ec) - ord(sc) + 1) * (el - sl + 1) // ine))
    return col, row, page


# 写入《南京科举博物馆进货明细》铅字补字表（多标签页）
# df: 输入的包含'A''B'两列的DataFrame
# f：要写入的EXCEL文件
# la：要写入的EXCEL文件的标签页
# sc：起始列
# ine：列间隔
# ec: 终止列（数字列）
# sl：起始行
# el：终止行

def writein(df, f, la, sc, ine, ec, sl, el):
    wb = load_workbook(f)
    # ws = wb[la]
    i = 1  # 标签页()内的数字，如：铅（金）字补字表 (2)
    for row in df.itertuples():
        # print(row[0], row[1], row[2])
        c, r, p = getxlspos(int(row[0]), sc, ine, ec, sl, el)
        print(c, r, p)
        ws = wb[la + ' (' + str(p) + ')' if p > 1 else la]
        ws[c + str(r)].value = row[1]
        ws[chr(ord(c) + 1) + str(r)].value = int(row[2])
        if int(row[3]) == 1000:
            # 3-设置样式，并且加载到对应单元格
            fill = PatternFill("solid", fgColor="EE7AE9")
            ws[c + str(r)].fill = fill
    wb.save(f)


# 对EXCEL表格中的数值进行整除
# f：要写入的EXCEL文件
# la：要写入的EXCEL文件的标签页
# sc：起始列
# ine：列间隔
# ec: 终止列（数字列）
# sl：起始行
# el：终止行
# d: 整除数（如2，3）
def divide(f, la, sc, ine, ec, sl, el):
    wb = load_workbook(f)
    ws = wb[la]
    for c in range(ord(sc), ord(ec) + 1, ine):
        col = chr(c)
        for row in range(sl, el + 1):
            val = ws[col + str(row)].value
            if val:
                ws[col + str(row)].value = math.ceil(val / 2)
                print(la, col + str(row), val, ws[col + str(row)].value)
    wb.save(f)


# df = sortposition('./11月补铅字.txt')
# print(getxlspos(47, 'A', 2, 7, 100))
# writein(df, './南京科举博物馆进货明细11.1.xlsx', '铅字补字表', 'A', 2, 7, 100)
# 铅字补字表 (2)
# print(getxlspos(1395, 'A', 2, 'N', 7, 99))
s = datetime.datetime.now()
# writein(sortposition('./12月待补铅字.txt'), '../字在进货/2019.12月进货/南京科举博物馆进货明细12.3.xlsx', '铅字补字表', 'A', 2, 'N', 7, 99)
# writein(sortposition('./2020年8月待补铅字.txt'), '../字在进货/2020.8月进货/字在科博店进货明细2020.8.5.xlsx', '活字补字表-一号字 2', 'A', 2, 'F', 7, 64)
# writein(sortposition('./2020年9月待补铅字.txt'), '../字在进货/2020.9月进货/字在科博店进货明细2020.9.1.xlsx', '活字补字表-一号字', 'A', 2, 'H', 7,
#        64)
#writein(sortposition('./2020年9月待补铅字.txt'), '../字在进货/2020.9月进货/字在科博店进货明细2020.9.10新.xlsx', '活字补字表-一号字', 'A', 2, 'L', 7,
#        64)
#writein(sortposition('./2020N南京字在10.3.txt'), '../字在进货/2020.11月进货/N南京字在10.3.xlsx', '铅字补字表（第三页）', 'A', 2, 'L', 7,
#        44)
#writein(sortposition('./2020南京西市补货单10.3.txt'), '../字在进货/2020.11月进货/南京西市补货单10.3.xlsx', '1号铅字补字表（第三页）', 'A', 2, 'L', 7,
#        44)
#writein(sortposition('./2020年11月待补铅字2.txt'), '../字在进货/2020.11月进货/南京科举博物馆进货明细11.18.xlsx', '铅字补字表（第三页）', 'A', 2, 'L', 7,
#        44)
#writein(sortposition('./2020年12月待补1号字另.txt'), './南京字在（明细）.xlsx', '1号宋繁', 'A', 2, 'L', 7, 40)
#writein(sortposition('./2020年12月待补3号字另.txt'), './南京字在（明细）.xlsx', '3号宋繁', 'A', 2, 'L', 7, 40)

#writein(sortposition('./2020年12月待补1号字.txt'), '../字在进货/2020.12月进货/南京科举博物馆进货明细12.27.xlsx', '铅字补字表 (1)', 'A', 2, 'L', 7,
#        99)
#writein(sortposition('./2020年12月待补3号字.txt'), '../字在进货/2020.12月进货/南京科举博物馆进货明细12.27.xlsx', '铅字补字表 (3)', 'A', 2, 'L', 7,
#        99)
#writein(sortposition('./2021年1月待补1号字.txt'), '../字在进货/2021.1月进货/南京科举博物馆进货明细2021.1.10 (1).xlsx', '铅字补字表 (1)', 'A', 2, 'L', 7,
#        99)
#writein(sortposition('./2021年1月待补3号字.txt'), '../字在进货/2021.1月进货/南京科举博物馆进货明细2021.1.10 (1).xlsx', '铅字补字表 (3)', 'A', 2, 'L', 7,
#        99)
#writein(sortposition('./2021年4月待补1号字.txt'), '../字在进货/2021.4月进货/补货单 2021年4月5日 南京市科博店.xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年4月待补3号字.txt'), '../字在进货/2021.4月进货/补货单 2021年4月5日 南京市科博店.xlsx',
#        '三号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./西市2021年1月待补1号字.txt'), '../其他门店/南京西市明细（活字） (1).xlsx', '1号宋繁', 'A', 2, 'L', 7, 40)
#writein(sortposition('./门东2021年1月待补1号字.txt'), '../其他门店/南京门东字在明细（活字） (1) .xlsx', '1号宋繁', 'A', 2, 'L', 7, 40)
#writein(sortposition('./门东2021年1月待补3号字.txt'), '../其他门店/南京门东字在明细（活字） (1) .xlsx', '3号宋繁', 'A', 2, 'L', 7, 40)

#writein(sortposition('./2021年4月门东待补铅1号字.txt'), './南京门东字在明细活字4.14.xlsx', '1号宋繁', 'A', 2, 'L', 7, 40)
#writein(sortposition('./2021年4月门东待补铅3号字.txt'), './南京门东字在明细活字4.14.xlsx', '3号宋繁', 'A', 2, 'L', 7, 40)
#writein(sortposition('./2021年4月西市待补铅1号字.txt'), './南京西市明细4.14.xlsx', '1号宋繁', 'A', 2, 'L', 7, 40)

#writein(sortposition('./2021年4月科博待补铅1号字.txt'), '../字在进货/2021.4月进货/补货单 2021年4月15日 南京市科博店 .xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年5月科博待补铅1号字.txt'), '../字在进货/2021.5月进货/补货单 2021年5月15日 南京市科博店.xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年5月科博待补铅1号字.txt'), '../字在进货/2021.5月进货/补货单 2021年5月15日 南京市科博店.xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年5月科博待补铅3号字.txt'), '../字在进货/2021.5月进货/补货单 2021年5月15日 南京市科博店.xlsx',
#        '三号宋繁补字表 ', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年5月西市待补铅1号字.txt'), './西市 0525.xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年6月科博待补铅1号字.txt'), '../字在进货/2021.6月进货/补货单 2021年6月1日 南京市科博店 .xlsx',
#        '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年6月科博待补铅3号字.txt'), '../字在进货/2021.6月进货/补货单 2021年6月1日 南京市科博店 .xlsx',
#        '三号宋繁补字表 ', 'A', 2, 'N', 7, 99)

#writein(sortposition('./2021年5月字在待补铅1号字.txt'), './字在 0515(1).xlsx', '一号宋繁补字表', 'A', 2, 'N', 7, 99)
#writein(sortposition('./2021年5月字在待补铅3号字.txt'), './字在 0515(1).xlsx', '三号宋繁补字表 ', 'A', 2, 'N', 7, 99)

# divide('南京科举博物馆进货明细12.3.xlsx','金字补字表','B',3,'L',7,44)
# divide('南京科举博物馆进货明细12.3.xlsx','金字补字表 (2)','B',3,'H',7,44)
# divide('南京科举博物馆进货明细12.3.xlsx','铅字补字表','B',2,'N',7,99)
# divide('南京科举博物馆进货明细12.3.xlsx','铅字补字表 (2)','B',2,'N',7,99)
# divide('南京科举博物馆进货明细12.3.xlsx','铅字补字表 (3)','B',2,'D',7,99)
# print(sortJinPosition('12月待补金字.txt'))
# writein(sortJinPosition('12月待补金字.txt'), '南京科举博物馆进货明细12.3.xlsx', '金字补字表', 'A', 3, 'L', 7, 44)

#writein(sortJinPosition('2021年4月科博待补金字.txt'), '../字在进货/2021.5月进货/补货单 2021年4月15日 南京市科博店 .xlsx',
#        '金字补字表', 'A', 3, 'L', 7, 44)
#writein(sortJinPosition('./2021年5月字在待补金字.txt'), './字在 0515(1).xlsx', '金字补字表', 'A', 2, 'L', 7, 44)

e = datetime.datetime.now()
print('All spended ' + str((e - s).seconds) + ' seconds.')
