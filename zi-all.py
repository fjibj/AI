# coding:utf-8
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import os,shutil
import os.path
import time
import sys
import re
from urllib import request, parse
import datetime


# 写入统计表EXCEL文件
# word_dict: 数据字典
# tf: 目标EXCEL文件名
# ta: 目标EXCEL标签页名
# tc: 目标列名（如J），注：只能写入一列
# isnum: 目标列是否为数值，默认True
# isout: 售出为True，补货为False（默认）
def writeIn(word_dict, tf, ta, tc, isout=False, isnum=True):
    s = datetime.datetime.now()
    wb = load_workbook(tf)
    ws = wb[ta]
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)
    font = Font(name='微软雅黑', charset=134, family=None, b=False, i=False, strike=None, outline=None, shadow=None, condense=None, color=None, extend=None, sz=12.0, u=None, vertAlign=None, scheme=None)
    fill = PatternFill("solid", fgColor="FFFF00") #设置单元格黄色背景
    for k, v in word_dict.items():
        b = 0  # 是否找到
        for i in range(2, maxrow + 1):
            c = ws['C%d' % i].value
            if c.strip() == k:
                ws['C%d' % i].font = font
                ws['D%d' % i].fill = fill
                if isnum:
                    ws[tc + '%d' % i].value = -int(v) if isout else int(v) 
                else:
                    ws[tc + '%d' % i].value = v
                b = 1
                break
        if b == 0:  # 未找到，添加到最后一行
            ws['C%d' % (maxrow + 1)].font = font
            ws['D%d' % (maxrow + 1)].fill = fill
            ws['C%d' % (maxrow + 1)] = k
            ws['D%d' % (maxrow + 1)].value = "=SUM(E%d:AK%d)" % (maxrow + 1, maxrow + 1)
            if isnum:
                ws[tc + '%d' % (maxrow + 1)].value = -int(v) if isout else int(v)
            else:
                ws[tc + '%d' % (maxrow + 1)].value = v
            #新增的字查找陈列盘号并写入B列
            b,l,c = getposition(k)
            if b!=1000:
                ws['B%d' % (maxrow + 1)].value = str(b)+'盘'
            maxrow += 1
            print("maxrow=%d" % maxrow)
    # 保存文件
    e0 = datetime.datetime.now()
    print('for spended '+str((e0 - s).seconds)+' seconds.')
    s1 = e0
    wb.save(tf)
    e1 = datetime.datetime.now()
    print('save spended '+str((e1 - s1).seconds)+' seconds.')
    e = datetime.datetime.now()
    print(tf+' saved! spended '+str((e - s).seconds)+' seconds.')


# 从补货单提取汉字和数量
# f: 补货单EXCEL文件名
# la: 补货单EXCEL标签页名
# sl: 补货单起始行号
# el: 补货单终止行号
# sc: 起始列
# ec: 终止列（数字列）
# st: 列步长（铅字为2，金字为3）
def getAdd(f, la, sl, el, sc, ec, st):
    s = datetime.datetime.now()
    word_dict = {}
    wb2 = load_workbook(f)
    ws2 = wb2[la]
    
    # 读取进货表汉字和数量
    z = 0
    for i in range(ord(sc), ord(ec), st):
        zc = chr(i)  # 字列
        xc = chr(i + 1)  # 数列
        # print(zc, xc)
        for j in range(sl, el + 1):
            p = ws2[zc + '%d' % j].value
            if p and len(p.strip()) == 1:
                if word_dict.__contains__(p.strip()):
                    word_dict[p.strip()] += int(ws2[xc + '%d' % j].value)
                else:
                    word_dict[p.strip()] = int(ws2[xc + '%d' % j].value)
                z += 1
    print(z)
    print(len(word_dict))
    e = datetime.datetime.now()
    print('获得补货数据,spended '+str((e - s).seconds)+' seconds.')
    return word_dict

# 统计铅字售出字频写入统计表
# f: 读入的文本文件
# tf: 目标EXCEL文件
# ta: 目标EXCEL文件标签页名
# tc: 目标列名（如J、K）
# isout: 售出为True，补货为False
def getOut(f):
    s = datetime.datetime.now()
    word_lst = []
    word_dict = {}
    exclude_str = " .，。！？、（）【】<>《》=：+-*—“”…0123456789"

    with open(f, "r", encoding='UTF-8') as fileIn:
        # 添加每一个字到列表中
        for line in fileIn:
            for char in line:
                if not char == '\xa0' and not char == '\n':
                    word_lst.append(char)

        # 用字典统计每个字出现的个数
        for char in word_lst:
            if char not in exclude_str:
                if char.strip() not in word_dict:  # strip去除各种空白
                    word_dict[char] = 1
                else:
                    word_dict[char] += 1

        print(len(word_dict))
        e = datetime.datetime.now()
        print('获得售出数据,spended '+str((e - s).seconds)+' seconds.')
        return word_dict

# 复制生成带时间戳的文件
def saveFileWithDate(f):
    tf = os.path.splitext(f)[0] + time.strftime('%Y%m%d', time.localtime()) + os.path.splitext(f)[1]
    shutil.copyfile(f, tf)
    print(tf+' saved.')

# 读库存盘文件获取{汉字:库存盘号}字典
# f: 读入的文本文件
# sp: 分隔符
def getKucunpan(f,sp):
    s = datetime.datetime.now()
    word_dict = {}
    with open(f, "r", encoding='UTF-8') as fileIn:
        # 添加每一个字到列表中
        for line in fileIn:
            c = line.split(sp)
            #print(c[0],c[1])
            for char in c[1]:
                if not char == '\xa0' and not char == '\n':
                    word_dict[char] = c[0]
    e = datetime.datetime.now()
    print('获得库存盘号数据,spended '+str((e - s).seconds)+' seconds.')
    return word_dict


#获得汉字对应的陈列盘位置
# hz: 某个汉字
def getposition(hz):
    s0 = datetime.datetime.now()
    s = parse.quote(hz)
    url = 'http://www.zizaicn.com/weixin/getwordpos_ver2.php?str=' + s
    f = request.urlopen(url)
    h = f.read().decode('utf-8')
    m = re.search(r'查询结果：</strong><br>(.*)<br></div>', h)
    m2 = re.match(r'(.*):(.*)盘 (.*)行 (.*)列', m.group(1))
    e = datetime.datetime.now()
    print('获取汉字陈列盘位置,spended '+str((e - s0).seconds)+' seconds.')
    if m2:
        return int(m2.group(2)), int(m2.group(3)), int(m2.group(4))
    else:
        return 1000, 1000, 1000

def getChenliepan(ws,i,tc):
    t = ws['C%d' % i].value
    b,l,c = getposition(t)
    if b!=1000:
        ws[tc+'%d' % i].value = str(b)+'盘'
    print(i,b,t)

#遍历统计表汉字写入其陈列盘
# tf: 目标EXCEL文件
# ta: 目标EXCEL文件标签页名
# tc: 陈列盘列名（如B）
def writeChenliepan(tf,ta,tc):
    s = datetime.datetime.now()
    wb = load_workbook(tf)
    ws = wb[ta]
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)

    for i in range(2, maxrow + 1):
        t = ws['C%d' % i].value
        b,l,c = getposition(t)
        if b!=1000:
            ws[tc+'%d' % i].value = str(b)+'盘'

    e0 = datetime.datetime.now()
    print('for spended '+str((e0 - s).seconds)+' seconds.')
    s1 = e0
    # 保存文件
    wb.save(tf)
    e1 = datetime.datetime.now()
    print('save spended '+str((e1 - s1).seconds)+' seconds.')
    e = datetime.datetime.now()
    print('刷新陈列盘数据：'+tf+' saved,spended '+str((e - s).seconds)+' seconds.')

##----------------------------------------------------------------------------------------------------------##

starttime = datetime.datetime.now()
# 铅字补货写入统计表
writeIn(getAdd('补货单20191105南京科举博物馆.xlsx', '铅字补字表', 7, 100, 'A', 'J', 2), '1_字盘统计表.xlsx', '字盘统计表', 'K')


# 统计铅字售出字频写入统计表
writeIn(getOut('11月售出.txt'), '1_字盘统计表.xlsx', '字盘统计表', 'J', isout=True)


# 写统计表库存盘
writeIn(getKucunpan('11月库存盘.txt','：'),'1_字盘统计表.xlsx','字盘统计表','A', isnum=False)


# 全量刷新统计表陈列盘
writeChenliepan('1_字盘统计表.xlsx','字盘统计表','B')


# 复制生成带时间戳的目标文件
saveFileWithDate('1_字盘统计表.xlsx')

endtime = datetime.datetime.now()
print('总共花费时间：'+str((endtime - starttime).seconds)+' seconds.')
