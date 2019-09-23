#coding:utf-8
from openpyxl import load_workbook

word_lst = []
word_dict = {}

exclude_str = " .，。！？、（）【】<>《》=：+-*—“”…0123456789" 

with open("9月销售字.txt","r",encoding='UTF-8') as fileIn:
    
    # 添加每一个字到列表中
    for line in fileIn:
        for char in line:
            word_lst.append(char)

    # 用字典统计每个字出现的个数       
    for char in word_lst:
        if char not in exclude_str:
            if char.strip() not in word_dict: # strip去除各种空白
                word_dict[char] = 1
            else :
                word_dict[char] += 1
	
	# 对word_dict中的每个key，在《1_字盘统计表.xlsx》中B列查找，如果找到，将-word_dict[key]（负的）写入找到行的EN（N为找到字所
	# 在的行）,若未找到该字，则在B列最后填加该字，并在同一行E列填入数字（负的）
    # 注：9月份在E列，10月份在F列，。。。。
    wb=load_workbook('1_字盘统计表.xlsx')
    ws = wb['字盘统计表']
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)
    for k,v in word_dict.items():
        b = 0	#是否找到
        for i in range(2,maxrow+1):
            c = ws['B%d' % i].value
            if c.strip() == k:
                ws['E%d' % i].value = -v
                b = 1
                break
        if b==0:	#未找到，添加到最后一行
            ws['B%d' % (maxrow+1)] = k
            ws['E%d' % (maxrow+1)].value = -v
            ws['C%d' % (maxrow+1)].value = "=SUM(D%d:R%d)" %(maxrow+1, maxrow+1)
            maxrow += 1
            print("maxrow=%d" % maxrow)
	# 保存文件
    wb.save('1_字盘统计表.xlsx')

		
			

