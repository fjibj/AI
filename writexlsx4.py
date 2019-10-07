#coding:utf-8
from openpyxl import load_workbook
'''
from win32com.client import Dispatch
 
def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()
'''
word_dict = {}
# 尧(03盘)9
'''
s = "风(31盘)10"
s1 = s[:s.find("(")]
s2 = s[s.find("(")+1:s.find(")")]
s3 = s[s.find(")")+1:]
k = [s2,s3]
print(k[0])
print(k[1])

'''
with open("10月剩余.txt","r",encoding='UTF-8') as fileIn:
    for line in fileIn:
        s1 = line[:line.find("(")].strip()
        s2 = line[line.find("(")+1:line.find(")")].strip()
        s3 = line[line.find(")")+1:].strip()
        word_dict[s1]=[s2,s3]
    #print(word_dict)
    
    # 写入字盘
    wb2=load_workbook('1_字盘统计表.xlsx', data_only=True)  #读取公式值，只读不保存
    ws2 = wb2['字盘统计表']
    wb=load_workbook('1_字盘统计表.xlsx')
    ws = wb['字盘统计表']
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)
    for k,v in word_dict.items():
        b = 0	#是否找到
        for i in range(2,maxrow+1):
            c = ws['C%d' % i].value
            if c.strip() == k:
                # 从ws2取出原库存数
                p = int(ws2['D%d' % i].value)
                # 写入10月份售出
                ws['H%d' % i].value = int(v[1]) - p
                # 写入陈列盘（木架）
                ws['B%d' % i] = v[0]
                b = 1
                #break
        if b==0:	#未找到，添加到最后一行
            ws['C%d' % (maxrow+1)] = k
            ws['H%d' % i].value = int(v[1])
            ws['B%d' % (maxrow+1)].value = v[0]
            ws['D%d' % (maxrow+1)].value = "=SUM(E%d:AK%d)" %(maxrow+1, maxrow+1)
            maxrow += 1
            print("maxrow=%d" % maxrow)
	# 保存文件
    wb.save('1_字盘统计表.xlsx')

    #just_open('1_字盘统计表.xlsx')