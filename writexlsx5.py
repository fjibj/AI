# coding:utf-8
from openpyxl import load_workbook

'''
word_dict = {}

with open("10月楼上金字剩余.txt","r",encoding='UTF-8') as fileIn:
    for line in fileIn:
        s = line.split(',')
        word_dict[s[0]]=s[1].strip()
    print(word_dict)

	# 楼上金字写入字盘
    wb=load_workbook('1_金字统计表.xlsx')
    ws = wb['字盘统计表']
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)
    for k,v in word_dict.items():
        b = 0	#是否找到
        for i in range(2,maxrow+1):
            c = ws['B%d' % i].value
            if c.strip() == k:
                ws['D%d' % i].value = int(v)
                b = 1
                break
        if b==0:	#未找到，添加到最后一行
            ws['B%d' % (maxrow+1)] = k
            ws['D%d' % (maxrow+1)].value = int(v)
            maxrow += 1
            print("maxrow=%d" % maxrow)
	# 保存文件
    wb.save('1_金字统计表.xlsx')
'''
'''
# 写入起始数量
# 加载《1_博物馆需要金字明细20190805.xls》
word_dict = {}
wb2 = load_workbook('1_博物馆需要金字明细20190805.xlsx')
ws2 = wb2['Sheet1']
for i in range(1,265):
	p = ws2['A%d' %i].value
	if p.strip():
		word_dict[p.strip()] = int(ws2['B%d' %i].value)

wb=load_workbook('1_金字统计表.xlsx')
ws = wb['字盘统计表']
maxrow = ws.max_row
print("maxrow=%d" % maxrow)

for k,v in word_dict.items():
	b = 0	#是否找到
	for i in range(2,maxrow+1):
		c = ws['B%d' % i].value
		if c.strip() == k:
			ws['E%d' % i].value = v
			b = 1
			break
	if b==0:	#未找到，添加到最后一行
		ws['B%d' % (maxrow+1)] = k
		ws['E%d' % (maxrow+1)].value = v
		maxrow += 1
		print("maxrow=%d" % maxrow)
# 保存文件
wb.save('1_金字统计表.xlsx')
'''
# 增加起始数量
# 加载《4_南京科举博物馆进货明细9.19-1.xlsx》
word_dict = {}
wb2 = load_workbook('3_9月南京科举博物馆进货明细2.xlsx')
# wb2 = load_workbook('南京科举博物馆进货明细9.19-1.xlsx')
ws2 = wb2['金字补字表']
for i in range(7, 45):
    p = ws2['A%d' % i].value
    if p.strip():
        word_dict[p.strip()] = int(ws2['B%d' % i].value)

for i in range(7, 45):
    p = ws2['D%d' % i].value
    if p.strip():
        word_dict[p.strip()] = int(ws2['E%d' % i].value)

for i in range(7, 45):
    p = ws2['G%d' % i].value
    if p and len(p.strip()) == 1:
        word_dict[p.strip()] = int(ws2['H%d' % i].value)

for i in range(7, 45):
    p = ws2['J%d' % i].value
    if p and len(p.strip()) == 1:
        word_dict[p.strip()] = int(ws2['K%d' % i].value)

print(len(word_dict))

wb = load_workbook('1_金字统计表.xlsx')
ws = wb['字盘统计表']
maxrow = ws.max_row
print("maxrow=%d" % maxrow)

for k, v in word_dict.items():
    b = 0  # 是否找到
    for i in range(2, maxrow + 1):
        c = ws['B%d' % i].value
        if c.strip() == k:
            p = int(ws['E%d' % i].value)
            ws['E%d' % i].value = int(v) + p
            b = 1
            break
    if b == 0:  # 未找到，添加到最后一行
        ws['B%d' % (maxrow + 1)] = k
        ws['E%d' % (maxrow + 1)].value = int(v)
        maxrow += 1
        print("maxrow=%d" % maxrow)
# 保存文件
wb.save('1_金字统计表.xlsx')
