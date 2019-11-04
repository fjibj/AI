#coding:utf-8
from openpyxl import load_workbook

word_dict = {}

# 加载《南京科举博物馆进货明细10.9-1.xlsx》
wb2 = load_workbook('南京科举博物馆进货明细10.9-1.xlsx')
ws2 = wb2['铅字补字表']
for i in range(7,45):
	p = ws2['A%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['B%d' %i].value)

for i in range(7,45):
	p = ws2['C%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['D%d' %i].value)

for i in range(7,45):
	p = ws2['E%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['F%d' %i].value)

for i in range(7,45):
	p = ws2['G%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['H%d' %i].value)

for i in range(7,100):
	p = ws2['I%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['J%d' %i].value)

for i in range(7,100):
	p = ws2['K%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['L%d' %i].value)

for i in range(7,100):
	p = ws2['M%d' %i].value
	if p and len(p.strip())==1:
		word_dict[p.strip()] = int(ws2['N%d' %i].value)

print(word_dict)

# 注：9月份售出在E列，9月份补字在F列，。。。。
wb=load_workbook('1_字盘统计表.xlsx')
ws = wb['字盘统计表']
maxrow = ws.max_row
print("maxrow=%d" % maxrow)


for k,v in word_dict.items():
	b = 0	#是否找到
	for i in range(2,maxrow+1):
		c = ws['C%d' % i].value
		if c.strip() == k:
			ws['I%d' % i].value = v
			b = 1
			break
	if b==0:	#未找到，添加到最后一行
		ws['C%d' % (maxrow+1)] = k
		ws['I%d' % (maxrow+1)].value = int(v)
		ws['D%d' % (maxrow+1)].value = "=SUM(E%d:AK%d)" %(maxrow+1, maxrow+1)
		maxrow += 1
		print("maxrow=%d" % maxrow)
# 保存文件
wb.save('1_字盘统计表.xlsx')
