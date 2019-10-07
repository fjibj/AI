#coding:utf-8
from openpyxl import load_workbook

word_dict = {}

with open("9月补字.txt","r",encoding='UTF-8') as fileIn:
    for line in fileIn:
        s = line.split(',')
        word_dict[s[0]]=s[1].strip()
    print(word_dict)

    # 写入字盘
    wb=load_workbook('1_字盘统计表.xlsx')
    ws = wb['字盘统计表']
    maxrow = ws.max_row
    print("maxrow=%d" % maxrow)
    for k,v in word_dict.items():
        b = 0	#是否找到
        for i in range(2,maxrow+1):
            c = ws['B%d' % i].value
            if c.strip() == k:
                p = ws['A%d' % i].value
                if p and p.strip():
                    ws['A%d' % i].value = v + '(原' + p + ')'
                else:
                    ws['A%d' % i].value = v
                b = 1
                break
        if b==0:	#未找到，添加到最后一行
            ws['B%d' % (maxrow+1)] = k
            ws['A%d' % (maxrow+1)].value = v
            maxrow += 1
            print("maxrow=%d" % maxrow)
	# 保存文件
    wb.save('1_字盘统计表.xlsx')
