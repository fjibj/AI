# coding:utf-8
import re
from urllib import request, parse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def getposition(hz):
    s = parse.quote(hz)
    url = 'http://www.zizaicn.com/weixin/getwordpos_ver2.php?str=' + s
    f = request.urlopen(url)
    h = f.read().decode('utf-8')
    m = re.search(r'查询结果：</strong><br>(.*)<br></div>', h)
    m2 = re.match(r'(.*):(.*)盘 (.*)行 (.*)列', m.group(1))
    if m2:
        return int(m2.group(2)), int(m2.group(3)), int(m2.group(4))
    else:
        return 1000, 1000, 1000


# print(getposition('面'))
# print(getposition('禀'))


# 方:21盘 4行 14列
# 堃:对不起,字库没有这个字

# 读.csv文件生成DataFrame
def sortposition(csvfile):
    df = pd.read_csv(csvfile, header=None, names=['A', 'B'])
    # print(df['A'])
    df[['C', 'D', 'E']] = df.apply(lambda row: pd.Series(getposition(row['A'])), axis=1)
    # print(df)
    # df.to_csv('sort1.csv', encoding='utf-8')
    df2 = df.sort_values(['C', 'D', 'E'], ascending=[True, True, True])
    df2.reset_index(drop=True, inplace=True)
    # df2.to_csv('sort2.csv', encoding='utf-8')
    return df2


# 获得在Excel文档中的行列位置
# index：在df中的序号
# sc：起始列
# ine：列间隔
# sl：起始行
# el：终止行
def getxlspos(index, sc, ine, sl, el):
    row = index % (el - sl + 1) + sl
    col = chr(ord(sc) + (index // (el - sl + 1)) * 2)
    return col, row


# 写入《南京科举博物馆进货明细》铅字补字表
# df: 输入的包含'A''B'两列的DataFrame
# f：要写入的EXCEL文件
# la：要写入的EXCEL文件的标签页
# sc：起始列
# ine：列间隔
# sl：起始行
# el：终止行

def writein(df, f, la, sc, ine, sl, el):
    wb = load_workbook(f)
    ws = wb[la]
    for row in df.itertuples():
        # print(row[0], row[1], row[2])
        c, r = getxlspos(int(row[0]), sc, ine, sl, el)
        ws[c + str(r)].value = row[1]
        ws[chr(ord(c) + 1) + str(r)].value = int(row[2])
        if int(row[3]) == 1000:
            # 3-设置样式，并且加载到对应单元格
            fill = PatternFill("solid", fgColor="EE7AE9")
            ws[c + str(r)].fill = fill
    wb.save(f)


df = sortposition('./11月补铅字.txt')
#print(getxlspos(47, 'A', 2, 7, 100))
writein(df, './南京科举博物馆进货明细11.1.xlsx', '铅字补字表', 'A', 2, 7, 100)
